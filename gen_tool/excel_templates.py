from __future__ import annotations

from dataclasses import dataclass
import copy
from pathlib import Path
from typing import Any

import json5
import openpyxl

import re

from gen_tool.constants import DEFAULT_TEMPLATE_XLSX, DIEUTIN_TYPES_ORDER, DieuTinType


@dataclass(frozen=True)
class TemplateBundle:
    by_type: dict[DieuTinType, dict[str, Any]]


def load_templates(xlsx_path: str | Path = DEFAULT_TEMPLATE_XLSX) -> TemplateBundle:
    p = Path(xlsx_path)
    wb = openpyxl.load_workbook(p, data_only=False)
    dtq_code_re = re.compile(r"^DTQ_(?P<code>[A-Za-z0-9]+)_\d+")

    by_type: dict[DieuTinType, dict[str, Any]] = {}

    def _parse_payload(ws_name: str) -> dict[str, Any]:
        ws = wb[ws_name]
        raw = ws["A1"].value
        if not isinstance(raw, str) or not raw.strip():
            raise ValueError(f"Sheet {ws_name} A1 is empty")
        payload = json5.loads(raw)
        if not isinstance(payload, dict):
            raise ValueError(f"Sheet {ws_name} A1 is not a JSON object")
        return payload

    for ws_name in wb.sheetnames:
        payload = _parse_payload(ws_name)

        code: str | None = None
        orders = payload.get("orders")
        if isinstance(orders, list) and orders and isinstance(orders[0], dict):
            order_id = str(orders[0].get("orderId", "")).strip()
            m = dtq_code_re.match(order_id)
            if m:
                code = m.group("code")

        if code is None:
            pickup_task_id = str(payload.get("pickupTaskId", "")).strip()
            m = dtq_code_re.match(pickup_task_id)
            if m:
                code = m.group("code")

        if code is None:
            continue

        dt = code
        if dt in by_type:
            existing_orders = by_type[dt].get("orders")
            existing_has_items = (
                isinstance(existing_orders, list)
                and existing_orders
                and isinstance(existing_orders[0], dict)
                and isinstance(existing_orders[0].get("items"), list)
                and bool(existing_orders[0].get("items"))
            )
            new_orders = payload.get("orders")
            new_has_items = (
                isinstance(new_orders, list)
                and new_orders
                and isinstance(new_orders[0], dict)
                and isinstance(new_orders[0].get("items"), list)
                and bool(new_orders[0].get("items"))
            )
            if existing_has_items and not new_has_items:
                continue

        by_type[dt] = payload

    if not by_type:
        legacy_map: dict[str, DieuTinType] = {
            "Lấy tổng": "BC",
            "Lấy từng đơn": "BC",
            "web api - có kiện": "WEB",
            "web api - ko kiện": "WEB",
        }
        for ws_name, dt in legacy_map.items():
            if ws_name not in wb.sheetnames:
                continue
            payload = _parse_payload(ws_name)
            if dt not in by_type:
                by_type[dt] = payload
            else:
                existing_orders = by_type[dt].get("orders")
                existing_has_items = (
                    isinstance(existing_orders, list)
                    and existing_orders
                    and isinstance(existing_orders[0], dict)
                    and isinstance(existing_orders[0].get("items"), list)
                    and bool(existing_orders[0].get("items"))
                )
                new_orders = payload.get("orders")
                new_has_items = (
                    isinstance(new_orders, list)
                    and new_orders
                    and isinstance(new_orders[0], dict)
                    and isinstance(new_orders[0].get("items"), list)
                    and bool(new_orders[0].get("items"))
                )
                if new_has_items:
                    by_type[dt] = payload

    if not by_type:
        raise ValueError(
            "Không tìm thấy template trong Excel. "
            "Hãy đảm bảo A1 chứa JSON và mã thuộc dạng DTQ_<LoaiDieuTin>_Index "
            "(ví dụ DTQ_WEB_0001)."
        )

    # Nếu Excel hiện tại chỉ có template cho 2 nhóm (vd: BC và WEB),
    # vẫn cần UI/logic hiển thị đủ 5 loại. Ta sẽ copy template gần nhất
    # để tạo đủ các key.
    src_any = next(iter(by_type.values()))
    src_bc = by_type.get("BC", src_any)
    src_web = by_type.get("WEB", src_any)
    defaults: dict[str, dict[str, Any]] = {
        "BC": src_bc,
        "HT": src_bc,
        "NH": src_web,
        "KL": src_web,
        "WEB": src_web,
    }
    for code in DIEUTIN_TYPES_ORDER:
        by_type.setdefault(code, copy.deepcopy(defaults[code]))

    return TemplateBundle(by_type=by_type)
